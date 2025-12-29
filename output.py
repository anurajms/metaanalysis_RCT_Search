"""
Output generation for Excel and CSV files.
"""

import csv
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import RCTRecord
from .utils import logger


# Column order for output
OUTPUT_COLUMNS = [
    'source_primary',
    'sources_found_in',
    'pmid',
    'pmcid',
    'doi',
    's2PaperId',
    'openalex_id',
    'title',
    'authors',
    'journal',
    'issn',
    'publication_date',
    'publication_year',
    'language',
    'abstract',
    'mesh_terms',
    'keywords',
    'fields_of_study',
    'url',
    'publisher',
    'rct_flag',
    'rct_detection_method',
    'topic',
    'classification_reason',
    'data_quality_notes',
]


def records_to_dataframe(records: List[RCTRecord]) -> pd.DataFrame:
    """Convert list of RCTRecords to a pandas DataFrame."""
    data = [record.to_dict() for record in records]
    df = pd.DataFrame(data)
    
    # Reorder columns
    existing_cols = [col for col in OUTPUT_COLUMNS if col in df.columns]
    extra_cols = [col for col in df.columns if col not in OUTPUT_COLUMNS]
    df = df[existing_cols + extra_cols]
    
    return df


def save_excel(records: List[RCTRecord], output_path: str) -> str:
    """
    Save records to an Excel file with formatting.
    
    Args:
        records: List of RCTRecord objects
        output_path: Path to output Excel file
        
    Returns:
        The actual path written to
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    df = records_to_dataframe(records)
    
    # Create workbook with openpyxl for formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "RCT Results"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Cell style
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                # Data row
                cell.alignment = cell_alignment
            
            cell.border = thin_border
    
    # Set column widths
    column_widths = {
        'source_primary': 12,
        'sources_found_in': 15,
        'pmid': 10,
        'pmcid': 12,
        'doi': 25,
        's2PaperId': 15,
        'openalex_id': 15,
        'title': 50,
        'authors': 30,
        'journal': 25,
        'issn': 12,
        'publication_date': 12,
        'publication_year': 8,
        'language': 8,
        'abstract': 60,
        'mesh_terms': 30,
        'keywords': 25,
        'fields_of_study': 25,
        'url': 30,
        'publisher': 20,
        'rct_flag': 8,
        'rct_detection_method': 35,
        'topic': 15,
        'classification_reason': 40,
        'data_quality_notes': 25,
    }
    
    for col_idx, col_name in enumerate(df.columns, 1):
        width = column_widths.get(col_name, 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    wb.save(output_path)
    logger.info(f"Saved Excel file: {output_path}")
    
    return str(output_path)


def save_csv(records: List[RCTRecord], output_path: str) -> str:
    """
    Save records to a CSV file.
    
    Args:
        records: List of RCTRecord objects
        output_path: Path to output CSV file
        
    Returns:
        The actual path written to
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    df = records_to_dataframe(records)
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    
    logger.info(f"Saved CSV file: {output_path}")
    return str(output_path)


def save_outputs(records: List[RCTRecord], excel_path: str) -> tuple:
    """
    Save records to both Excel and CSV files.
    
    Args:
        records: List of RCTRecord objects
        excel_path: Path to output Excel file (CSV will be same name with .csv extension)
        
    Returns:
        Tuple of (excel_path, csv_path)
    """
    excel_path = Path(excel_path)
    csv_path = excel_path.with_suffix('.csv')
    
    excel_result = save_excel(records, str(excel_path))
    csv_result = save_csv(records, str(csv_path))
    
    return excel_result, csv_result
